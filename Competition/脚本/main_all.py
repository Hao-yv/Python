import gradio as gr
import os
import openpyxl
import info
from tqdm import tqdm
import time
import requests
import re
from bs4 import BeautifulSoup
from urllib.parse import urlparse
configData = []


sheet_head = ['项目编号', '项目名称', '项目类型', '所属学校', '项目期限', '所属一级学科',
              '所属二级学科', '立项时间', '结题时间', '指导老师', '职称', '指导老师类型', '负责人', '其他成员']

# 定义可接受参数的main函数


def maining(urlPrefix, LXYear, SchoolName, progress=gr.Progress()):
    progress(0,desc="开始...")
    urlList = info.getUrlList(
        '{}/Index/ItemPageList?LXYear={}&SchoolName={}'.format(urlPrefix, LXYear, SchoolName), progress)
    print('获取urlList成功')
    return (urlList)


def down(LXYear, SchoolName, urlList, progress=gr.Progress()):
    urlList = eval(urlList)
    print('开始下载')
    dataArr = []
    for url in progress.tqdm(urlList):
        # time.sleep(0.1)
        dataArr.append(info.getInfo(url))
    wbResult = openpyxl.Workbook()
    wsR = wbResult.active
    wsR.append(sheet_head)
    print('正在导出为excel')
    for row in progress.tqdm(dataArr):
        # time.sleep(0.1)
        wsR.append(row)
    # 保存到/excel目录下
    if not os.path.exists('excel'):
        os.mkdir('excel')
    targetFileName = '{}.xlsx'.format(SchoolName+LXYear)
    try:
        wbResult.save('excel/'+targetFileName)
    except PermissionError:
        return ('要保存的excel被其他程序占用')
    print('数据获取完成')
    return ('数据获取完成,保存到{}\{}'.format(os.getcwd(), targetFileName))



title = '大创webUI'

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}


def getUrlList(url, progress):
    firstRes = requests.get(url, headers=headers)
    pages = 0
    # 取class=pager-info
    if firstRes.status_code == 200:
        soup = BeautifulSoup(firstRes.text, 'lxml')
        pages = soup.find('div', class_='pager-info').text
        reMatch = re.search(r'共(\d+)页(\d+)条记录', pages)
        pages = int(reMatch.group(1))
        records = int(reMatch.group(2))
        print('符合条件的记录有{}页,{}条'.format(pages, records))

    pageUrlList = []
    infoUrlList = []
    for pageIndex in range(1, pages+1):
        pageUrlList.append(url+"&PageIndex="+str(pageIndex))
    for pUrl in progress.tqdm(pageUrlList):
        time.sleep(0.1)
        res = requests.get(pUrl, headers=headers)
        # 取class=btn-group btn-group-sm的href
        if res.status_code == 200:
            soup = BeautifulSoup(res.text, 'lxml')
            for item in soup.find_all('a', class_='btn-sm'):
                # url 只取域名
                domain = 'http://'+urlparse(url).netloc
                infoUrlList.append(domain+item['href'])
    print(infoUrlList, 'infoList')
    return infoUrlList


def getInfo(url):
    '''
    # 初始化一个对象
    project = Project()
    project.项目名称 = '1230'
    print(project)
    '''
    r = requests.get(url, headers=headers)
    row = []
    stuRow = []
    teacherRow = []
    # 取class=c-detail-item的元素,排除
    if r.status_code == 200:
        soup = BeautifulSoup(r.text, 'lxml')
        for idx, val in enumerate(soup.find_all('div', class_='form-group-line')):
            # 子元素class=c-label-title的text作为键，子元素class=c-detail-item的text作为值,去掉\n和\r
            for i in val.find_all('div', class_='c-detail-item'):
                row.append(i.text.strip().replace(
                    '\r\n                                    ', ''))
        # 取table标签
        for idx, val in enumerate(soup.find_all('table')):
            if idx == 0:
                for i in val.find_all('tr'):
                    arr = []
                    for j in i.find_all('td'):
                        arr.append(j.text.strip())
                    if arr:
                        stuRow.append(arr)
            elif idx == 1:
                for i in val.find_all('tr'):
                    arr = []
                    for j in i.find_all('td'):
                        arr.append(j.text.strip())
                    if arr:
                        teacherRow.append(arr)

    # print(row)
    # print(stuRow, '学生')
    # print(teacherRow, '教师')
    # teacherRow的数组合并，如[['1','2','3'],['4','5','6']]合并为['1,4','2,5','3,6']
    merged = [','.join(items) for items in zip(*teacherRow)]
    row.extend(merged)
    # print(merged, '合并后教师')
    # print(row, '合并教师后数据')
    # stuRow中取出第一个数组的第一个元素，并删除
    row.append(stuRow.pop(0)[0])
    # 合并学生数组
    merged = [','.join(items) for items in zip(*stuRow)]
    row.extend(merged)
    # print(row, '合并学生后数据')
    return row


# if __name__ == '__main__':
    # getInfo('http://cxcy.sdei.edu.cn/cxxl/cxxl/Index/ItemDetail?id=e6051b14-031f-410e-9aeb-4381296667db')
    # getUrlList('http://cxcy.sdei.edu.cn/cxxl/cxxl/Index/ItemPageList?LXYear=2022&SchoolName=潍坊学院')




def getConfigDataValue(key):
    return configData[key]


if os.path.exists('./config.xlsx'):
    print('配置文件config.xlsx加载中')
    wb = openpyxl.load_workbook('./config.xlsx')
    ws = wb.active
    for row in ws.iter_rows():
        iRow = []
        for cell in row:
            iRow.append(cell.value)
        configData.append(iRow)
    print('配置文件config.xlsx加载成功')
    print('正在启动大创webUI')

# Url
urlText = gr.Text(label='官网地址')
# Result
resultText = gr.Textbox(label='结果')
# print(configData)
# 将二维数组data转为字典
configData = dict(configData)
# print(configData, '转化字典')
urlDropDown = gr.Dropdown(
    label='请选择省份', choices=configData)

# urlDropDown.change(fn=getConfigDataValue,inputs=urlDropDown,outputs=urlText)
with gr.Blocks() as demo:
    gr.Markdown('# 大创 webUI')
    gr.Markdown('#### by leev')
    urlDropDown = gr.Dropdown(
        label='请选择省份', choices=configData)
    urlText = gr.Textbox(label='官网地址', interactive=False)
    urlDropDown.change(fn=getConfigDataValue,
                       inputs=urlDropDown, outputs=urlText)
    # 年份
    inputYear = gr.Text(type='text', label='请输入年份')
    # 学校
    inputSchool = gr.Text(type='text', label='请输入学校')
    greet_btn = gr.Button("搜索", variant='primary')
    out = gr.Textbox(label='搜索结果', lines=3)
    greet_btn.click(fn=maining, inputs=[
                    urlText, inputYear, inputSchool], outputs=out)
    down_btn = gr.Button("下载", variant='primary')
    result = gr.Textbox(label='下载结果', lines=1)
    down_btn.click(fn=down, inputs=[
                   inputYear, inputSchool, out], outputs=result)
# demo = gr.Interface(theme=gr.themes.Base().background_fill_primary_dark, title=title, fn=getConfigDataValue,inputs=inputs, outputs=resultText)
print('启动成功,请不要关闭此窗口')
demo.queue().launch(server_port=30002, inbrowser=True)


