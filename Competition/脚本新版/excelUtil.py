import openpyxl
import os

fileName = str(input("请输入文件名："))

outFileName = str(input("请输入要转换的文件名"))
# 读取excel目录下所有的xlsx文件
path = os.getcwd()
print('读取文件为:'+path + '\\{}.xlsx'.format(fileName))
wb1 = openpyxl.load_workbook(path + '\excel\{}.xlsx'.format(fileName))
ws1 = wb1.active

# 读取当前目录下的江苏省高等学校大学生创新创业训练计划2021年立项项目名单.xlsx
wb2 = openpyxl.load_workbook(path + '\\{}.xlsx'.format(outFileName))
ws2 = wb2.active

table1 = []
table2 = []

for row in ws1.iter_rows():
    tRow = []
    for cell in row:
        tRow.append(cell.value)
    table1.append(tRow)

for row in ws2.iter_rows():
    tRow = []
    for cell in row:
        tRow.append(cell.value)
    table2.append(tRow)

dict1 = {item[1]: item for item in table1}

# 循环table2
for i in range(len(table2)):
    key = table2[i][6]
    if key in dict1:
        table2[i][8] = dict1[key][5]
        table2[i][9] = dict1[key][6]
        table2[i][3] = dict1[key][3]


wb = openpyxl.Workbook()
ws = wb.active
for i in range(len(table2)):
    ws.append(table2[i])
wb.save('{}.xlsx'.format(outFileName))
print('done')
